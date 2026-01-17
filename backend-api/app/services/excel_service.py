"""
엑셀 및 CSV 파일 처리 서비스
"""
from io import BytesIO, StringIO
from typing import List, Dict, Any
from datetime import datetime, date
from decimal import Decimal
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.models import Transaction, Category


def export_transactions_to_excel(
    db: Session,
    transactions: List[Transaction],
    categories: Dict[int, Category]
) -> BytesIO:
    """
    거래 내역을 엑셀 파일로 변환
    
    Args:
        db: 데이터베이스 세션
        transactions: 거래 내역 리스트
        categories: 카테고리 딕셔너리 (category_id -> Category)
    
    Returns:
        BytesIO: 엑셀 파일 바이트 스트림
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "거래 내역"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    headers = ["날짜", "유형", "카테고리", "금액", "설명", "등록일시", "수정일시"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 데이터 작성
    for row_idx, transaction in enumerate(transactions, start=2):
        category = categories.get(transaction.category_id)
        category_name = category.name if category else "알 수 없음"
        
        # 날짜 형식 변환
        transaction_date = transaction.transaction_date
        if isinstance(transaction_date, str):
            transaction_date = datetime.strptime(transaction_date, "%Y-%m-%d").date()
        
        ws.cell(row=row_idx, column=1, value=transaction_date)
        ws.cell(row=row_idx, column=2, value="수입" if transaction.type == "income" else "지출")
        ws.cell(row=row_idx, column=3, value=category_name)
        ws.cell(row=row_idx, column=4, value=float(transaction.amount))
        ws.cell(row=row_idx, column=5, value=transaction.description or "")
        
        # 날짜/시간 형식 변환
        created_at = transaction.created_at
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        ws.cell(row=row_idx, column=6, value=created_at)
        
        updated_at = transaction.updated_at
        if isinstance(updated_at, str):
            updated_at = datetime.fromisoformat(updated_at.replace("Z", "+00:00"))
        ws.cell(row=row_idx, column=7, value=updated_at)
    
    # 열 너비 자동 조정
    column_widths = {
        "A": 12,  # 날짜
        "B": 10,  # 유형
        "C": 20,  # 카테고리
        "D": 15,  # 금액
        "E": 30,  # 설명
        "F": 20,  # 등록일시
        "G": 20,  # 수정일시
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 금액 열 숫자 형식 설정
    for row in range(2, len(transactions) + 2):
        cell = ws.cell(row=row, column=4)
        cell.number_format = "#,##0"
    
    # 날짜 열 형식 설정
    for row in range(2, len(transactions) + 2):
        ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
        ws.cell(row=row, column=6).number_format = "YYYY-MM-DD HH:MM:SS"
        ws.cell(row=row, column=7).number_format = "YYYY-MM-DD HH:MM:SS"
    
    # 파일을 메모리에 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def import_transactions_from_excel(
    db: Session,
    file_content: bytes,
    user_id: int,
    categories: Dict[str, Category]  # category_name -> Category
) -> Dict[str, Any]:
    """
    엑셀 파일에서 거래 내역을 읽어서 데이터베이스에 저장
    
    Args:
        db: 데이터베이스 세션
        file_content: 엑셀 파일 바이트
        user_id: 사용자 ID
        categories: 카테고리 딕셔너리 (category_name -> Category)
    
    Returns:
        Dict: {"success": int, "failed": int, "errors": List[str]}
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(BytesIO(file_content), data_only=True)
    ws = wb.active
    
    success_count = 0
    failed_count = 0
    errors = []
    
    # 헤더 행 건너뛰기 (1행)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # 빈 행 건너뛰기
            if not row[0]:  # 날짜가 없으면 빈 행
                continue
            
            # 데이터 파싱
            transaction_date = row[0]
            if isinstance(transaction_date, datetime):
                transaction_date = transaction_date.date()
            elif isinstance(transaction_date, str):
                transaction_date = datetime.strptime(transaction_date, "%Y-%m-%d").date()
            
            type_str = str(row[1]).strip()
            if type_str not in ["수입", "지출", "income", "expense"]:
                raise ValueError(f"유형이 올바르지 않습니다: {type_str}")
            transaction_type = "income" if type_str in ["수입", "income"] else "expense"
            
            category_name = str(row[2]).strip()
            if not category_name:
                raise ValueError("카테고리명이 비어있습니다")
            
            # 카테고리 찾기 또는 생성
            if category_name not in categories:
                # 카테고리가 없으면 생성 (기본 색상 사용)
                category = Category(
                    name=category_name,
                    type=transaction_type,
                    user_id=user_id,
                    color="#6b7280"
                )
                db.add(category)
                db.flush()  # ID를 얻기 위해 flush
                categories[category_name] = category
            else:
                category = categories[category_name]
                # 카테고리 타입이 일치하는지 확인
                if category.type != transaction_type:
                    raise ValueError(
                        f"카테고리 '{category_name}'의 타입({category.type})이 거래 유형({transaction_type})과 일치하지 않습니다"
                    )
            
            amount = float(row[3])
            if amount < 0:
                raise ValueError(f"금액은 0원 이상이어야 합니다: {amount}")
            
            description = str(row[4]).strip() if row[4] else None
            
            # 거래 내역 생성
            transaction = Transaction(
                user_id=user_id,
                category_id=category.id,
                type=transaction_type,
                amount=Decimal(str(amount)),
                description=description,
                transaction_date=transaction_date
            )
            db.add(transaction)
            success_count += 1
            
        except Exception as e:
            failed_count += 1
            error_msg = f"행 {row_idx}: {str(e)}"
            errors.append(error_msg)
            continue
    
    # 변경사항 저장
    if success_count > 0:
        db.commit()
    
    return {
        "success": success_count,
        "failed": failed_count,
        "errors": errors
    }


def export_categories_to_excel(
    db: Session,
    categories: List[Category]
) -> BytesIO:
    """
    카테고리를 엑셀 파일로 변환
    
    Args:
        db: 데이터베이스 세션
        categories: 카테고리 리스트
    
    Returns:
        BytesIO: 엑셀 파일 바이트 스트림
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "카테고리"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    headers = ["이름", "유형", "색상", "아이콘", "등록일시", "수정일시"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 데이터 작성
    for row_idx, category in enumerate(categories, start=2):
        ws.cell(row=row_idx, column=1, value=category.name)
        ws.cell(row=row_idx, column=2, value="수입" if category.type == "income" else "지출")
        ws.cell(row=row_idx, column=3, value=category.color or "")
        ws.cell(row=row_idx, column=4, value=category.icon or "")
        
        # 날짜/시간 형식 변환
        created_at = category.created_at
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        ws.cell(row=row_idx, column=5, value=created_at)
        
        updated_at = category.updated_at
        if isinstance(updated_at, str):
            updated_at = datetime.fromisoformat(updated_at.replace("Z", "+00:00"))
        ws.cell(row=row_idx, column=6, value=updated_at)
    
    # 열 너비 자동 조정
    column_widths = {
        "A": 20,  # 이름
        "B": 10,  # 유형
        "C": 12,  # 색상
        "D": 15,  # 아이콘
        "E": 20,  # 등록일시
        "F": 20,  # 수정일시
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 날짜 열 형식 설정
    for row in range(2, len(categories) + 2):
        ws.cell(row=row, column=5).number_format = "YYYY-MM-DD HH:MM:SS"
        ws.cell(row=row, column=6).number_format = "YYYY-MM-DD HH:MM:SS"
    
    # 파일을 메모리에 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def import_categories_from_excel(
    db: Session,
    file_content: bytes,
    user_id: int
) -> Dict[str, Any]:
    """
    엑셀 파일에서 카테고리를 읽어서 데이터베이스에 저장
    
    Args:
        db: 데이터베이스 세션
        file_content: 엑셀 파일 바이트
        user_id: 사용자 ID
    
    Returns:
        Dict: {"success": int, "failed": int, "errors": List[str]}
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(BytesIO(file_content), data_only=True)
    ws = wb.active
    
    success_count = 0
    failed_count = 0
    errors = []
    
    # 헤더 행 건너뛰기 (1행)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # 빈 행 건너뛰기
            if not row[0]:  # 이름이 없으면 빈 행
                continue
            
            # 데이터 파싱
            name = str(row[0]).strip()
            if not name:
                raise ValueError("카테고리 이름이 비어있습니다")
            
            type_str = str(row[1]).strip()
            if type_str not in ["수입", "지출", "income", "expense"]:
                raise ValueError(f"유형이 올바르지 않습니다: {type_str}")
            category_type = "income" if type_str in ["수입", "income"] else "expense"
            
            color = str(row[2]).strip() if row[2] else None
            if color and not color.startswith("#"):
                # #이 없으면 추가
                if len(color) == 6:
                    color = "#" + color
                else:
                    color = None
            
            icon = str(row[3]).strip() if row[3] else None
            
            # 기존 카테고리 확인 (같은 이름, 같은 유형)
            existing = db.query(Category).filter(
                Category.name == name,
                Category.type == category_type,
                Category.user_id == user_id
            ).first()
            
            if existing:
                # 기존 카테고리 업데이트
                if color:
                    existing.color = color
                if icon:
                    existing.icon = icon
                success_count += 1
            else:
                # 새 카테고리 생성
                category = Category(
                    name=name,
                    type=category_type,
                    user_id=user_id,
                    color=color or "#6b7280",
                    icon=icon
                )
                db.add(category)
                success_count += 1
            
        except Exception as e:
            failed_count += 1
            error_msg = f"행 {row_idx}: {str(e)}"
            errors.append(error_msg)
            continue
    
    # 변경사항 저장
    if success_count > 0:
        db.commit()
    
    return {
        "success": success_count,
        "failed": failed_count,
        "errors": errors
    }
